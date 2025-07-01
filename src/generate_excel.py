"""
title: Generate Excel Document
author: openlab
version: 0.1
description: Génère un fichier Excel via un LLM (Ollama) et renvoie un lien de téléchargement
"""

import os, uuid
from typing import Optional, Callable, Any, List, Dict
from pathlib import Path
from fastapi import UploadFile, Request
import re
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from open_webui.routers.files import upload_file
from open_webui.models.users import Users
from open_webui.storage.provider import Storage
from open_webui.models.files import Files, FileForm

class EventEmitter:
    def __init__(self, event_emitter: Callable[[dict], Any] = None):
        self.event_emitter = event_emitter

    async def emit(self, description="Unknown State", status="in_progress", done=False):
        if self.event_emitter:
            await self.event_emitter(
                {
                    "type": "status",
                    "data": {
                        "status": status,
                        "description": description,
                        "done": done,
                    },
                }
            )

class HelpFunctions:
    def __init__(self):
        self.default_fonts = {
            "main": "Calibri",
            "header": "Arial",
        }
        self.colors = {
            "header_bg": "1F4E78",
            "alt_row": "F2F2F2",
            "grid": "D3D3D3"
        }

    def format_header_row(self, worksheet, row_num, columns):
        """
        Format the header row with styling.
        
        Args:
            worksheet: The Excel worksheet.
            row_num (int): The row number for the header.
            columns (list): List of column names.
        """
        # Define styles
        header_font = Font(name=self.default_fonts["header"], bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=self.colors["header_bg"], end_color=self.colors["header_bg"], fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Apply styles to header row
        for col_idx, column_name in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_idx)
            cell.value = column_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
            # Adjust column width based on content
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = max(12, len(column_name) + 2)

    def format_data_rows(self, worksheet, start_row, end_row, num_cols):
        """
        Format the data rows with alternating colors and borders.
        
        Args:
            worksheet: The Excel worksheet.
            start_row (int): Starting row number for data.
            end_row (int): Ending row number for data.
            num_cols (int): Number of columns.
        """
        # Define styles
        alt_fill = PatternFill(start_color=self.colors["alt_row"], end_color=self.colors["alt_row"], fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color=self.colors["grid"]),
            right=Side(style='thin', color=self.colors["grid"]),
            top=Side(style='thin', color=self.colors["grid"]),
            bottom=Side(style='thin', color=self.colors["grid"])
        )
        
        # Apply styles to data rows
        for row in range(start_row, end_row + 1):
            # Apply alternating row colors
            if row % 2 == 0:
                for col in range(1, num_cols + 1):
                    worksheet.cell(row=row, column=col).fill = alt_fill
            
            # Apply borders to all cells
            for col in range(1, num_cols + 1):
                worksheet.cell(row=row, column=col).border = thin_border
                worksheet.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center")

    def create_excel_table(self, worksheet, start_row, end_row, num_cols, table_name):
        """
        Create an Excel table from the data range.
        
        Args:
            worksheet: The Excel worksheet.
            start_row (int): Starting row number for table.
            end_row (int): Ending row number for table.
            num_cols (int): Number of columns.
            table_name (str): Name for the table.
        """
        # Define table range
        table_ref = f"A{start_row}:{get_column_letter(num_cols)}{end_row}"
        
        # Create table
        table = Table(displayName=table_name, ref=table_ref)
        
        # Add a default style
        style = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        # Add the table to the worksheet
        worksheet.add_table(table)

    def format_worksheet(self, worksheet, title):
        """
        Apply general formatting to the worksheet.
        
        Args:
            worksheet: The Excel worksheet.
            title (str): Title for the worksheet.
        """
        # Set worksheet title
        worksheet.title = title[:31]  # Excel limits worksheet names to 31 chars
        
        # Add title at the top
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(size=14, bold=True)
        
        # Merge cells for title
        worksheet.merge_cells('A1:D1')
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add some space below title
        worksheet.row_dimensions[2].height = 10

# --- Tools ---
class Tools:
    def __init__(self):
        self.FILES_DIR = "./tmp"
        self.API_BASE_URL = "http://localhost:3000/api/v1/files/"
        os.makedirs(self.FILES_DIR, exist_ok=True)
        self.help_functions = HelpFunctions()
        self.event_emitter = EventEmitter()
    
    async def generate_excel_from_json(self, json_data: dict, __request__: Request, __event_emitter__: Callable[[dict], Any] = None, __user__=None):
        """
        Generate an Excel document from a JSON file.
        
        Args:
            json_data: The JSON data to generate the document from.
            __user__: The user to upload the file to.
        Returns:
            str: The download URL of the uploaded file.
        
        Example:
        Here is an example of the JSON data for a complete Excel document:
            
        ```json
        {
            "titre": "Rapport financier",
            "feuilles": [
                {
                    "nom": "Résumé",
                    "tableau": {
                        "colonnes": ["Mois", "Revenus", "Dépenses", "Profit"],
                        "données": [
                            ["Janvier", 10000, 8000, 2000],
                            ["Février", 12000, 7500, 4500],
                            ["Mars", 15000, 9000, 6000]
                        ]
                    }
                },
                {
                    "nom": "Détails dépenses",
                    "tableau": {
                        "colonnes": ["Catégorie", "Montant", "Pourcentage"],
                        "données": [
                            ["Salaires", 5000, "60%"],
                            ["Loyer", 1500, "20%"],
                            ["Fournitures", 800, "10%"],
                            ["Marketing", 500, "5%"],
                            ["Divers", 400, "5%"]
                        ]
                    }
                }
            ]
        }
        ```
        """
        emitter = EventEmitter(__event_emitter__)
        print("[DEBUG] json_data", json_data)
        titre = json_data.get('titre')
        print("[DEBUG] titre", titre)
        await emitter.emit(f"Initiating Excel generation for: {titre}")
        
        # Create workbook
        wb = Workbook()
        print("[DEBUG] workbook created")
        
        # Remove default worksheet if we have sheets defined
        if json_data.get('feuilles') and len(json_data.get('feuilles')) > 0:
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
        
        # Process each sheet
        try:
            await emitter.emit("Creating Excel sheets and tables")
            
            for sheet_idx, sheet_data in enumerate(json_data.get('feuilles', [])):
                sheet_name = sheet_data.get('nom', f'Feuille {sheet_idx+1}')
                
                # Create new worksheet
                if sheet_idx == 0 and "Sheet" in wb.sheetnames:
                    # Use default sheet for first worksheet
                    ws = wb.active
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(title=sheet_name)
                
                print(f"[DEBUG] Created sheet: {sheet_name}")
                
                # Format worksheet with title
                self.help_functions.format_worksheet(ws, sheet_data.get('nom', sheet_name))
                
                # Process table data
                if 'tableau' in sheet_data:
                    # Get table data
                    table_data = sheet_data['tableau']
                    columns = table_data.get('colonnes', [])
                    data_rows = table_data.get('données', [])
                    
                    # Start position (after title and spacing)
                    header_row = 3
                    
                    # Add header
                    self.help_functions.format_header_row(ws, header_row, columns)
                    
                    # Add data
                    for row_idx, row_data in enumerate(data_rows, header_row + 1):
                        for col_idx, cell_value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx).value = cell_value
                    
                    # Format data rows
                    if data_rows:
                        end_row = header_row + len(data_rows)
                        self.help_functions.format_data_rows(ws, header_row + 1, end_row, len(columns))
                        
                        # Create Excel table
                        table_name = f"Table{sheet_idx}".replace(" ", "")
                        try:
                            self.help_functions.create_excel_table(ws, header_row, end_row, len(columns), table_name)
                        except Exception as e:
                            print(f"[DEBUG] Could not create table: {e}")
                
                # Add simple formulas if it's a numeric table
                if 'tableau' in sheet_data:
                    columns = sheet_data['tableau'].get('colonnes', [])
                    data_rows = sheet_data['tableau'].get('données', [])
                    
                    if data_rows and len(data_rows) > 0:
                        numeric_cols = []
                        
                        # Detect numeric columns 
                        for col_idx, _ in enumerate(columns, 1):
                            if all(isinstance(row[col_idx-1], (int, float)) for row in data_rows if col_idx-1 < len(row)):
                                numeric_cols.append(col_idx)
                        
                        # Add totals row with formulas
                        if numeric_cols:
                            total_row = header_row + len(data_rows) + 1
                            ws.cell(row=total_row, column=1).value = "Total"
                            ws.cell(row=total_row, column=1).font = Font(bold=True)
                            
                            for col_idx in numeric_cols:
                                col_letter = get_column_letter(col_idx)
                                start_row = header_row + 1
                                end_row = header_row + len(data_rows)
                                formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
                                ws.cell(row=total_row, column=col_idx).value = formula
                                ws.cell(row=total_row, column=col_idx).font = Font(bold=True)
            
            await emitter.emit(
                status="complete",
                description=f"Excel generation completed",
                done=True,
            )
        except Exception as e:
            print("[DEBUG] Error", e)
            return f"Error: {str(e)}"
        
        # Save workbook
        if not os.path.exists(self.FILES_DIR):
            os.makedirs(self.FILES_DIR)
        # clean up title for filename
        clean_title = re.sub(r'[^\w\s]', '', json_data.get('titre', 'excel'))
        clean_title = clean_title.replace(' ', '_')

        output_path = self.FILES_DIR + '/' + clean_title + '.xlsx'
        wb.save(output_path)
        print("[DEBUG] output_path", output_path)

        try:
            with open(output_path, 'rb') as f:
                print("[DEBUG] f", f)
                files = UploadFile(file=f, filename=os.path.basename(output_path))
                print("[DEBUG] files", files)
                file_item = await self.upload_file(file=files, user_id=__user__['id'], __request__=__request__, __user__=__user__, __event_emitter__=__event_emitter__)
                print("[DEBUG] file_item", file_item)
                return file_item
        except Exception as e:
            print("[DEBUG] Error", e)
            return f"Error: {str(e)}"

    async def upload_file(self, file: UploadFile, user_id: str, __request__: Request, __user__: dict, __event_emitter__: Callable[[dict], Any] = None):
        emitter = EventEmitter(__event_emitter__)
        metadata = {"data": {"generated_by": "upload_file"}}
 
        await emitter.emit(f"Getting download link for file: {file.filename}")
        
        # get the user for permissions
        user = Users.get_user_by_id(id=__user__['id'])
        print("[DEBUG] user", user)
        # upload the file to the database
        doc = upload_file(request=__request__, file=file, user=user, metadata=metadata, process=False)
        print("[DEBUG] doc", doc)

        # get the download link
        download_link = f"{self.API_BASE_URL}{doc.id}/content"
        print("[DEBUG] download_link", download_link)
        await emitter.emit(
                status="complete",
                description=f"Finished generating the Excel file",
                done=True
            )
        return (
            f"<source><source_id>{doc.filename}</source_id><source_context>" 
            + str(download_link)
            + "</source_context></source>\n"
        )
